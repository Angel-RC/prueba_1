from __future__ import annotations

import email
import io
from email.header import decode_header
from email.utils import getaddresses, parsedate_to_datetime
from pathlib import Path

import openpyxl
import pdfplumber

from models.attachment import Attachment, ExcelInfo, ExcelLLMInfo, PDFInfo
from models.email import Email, EmailMetadata
from parser.excel import get_processor


class EmlParser:
    def parse(self, path: Path) -> Email:
        raw = path.read_bytes()
        msg = email.message_from_bytes(raw)

        metadata = self._extract_metadata(msg)
        attachments = self._extract_attachments(msg)

        return Email(metadata=metadata, attachments=attachments, source_path=str(path))

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _extract_metadata(self, msg) -> EmailMetadata:
        date = None
        if msg["Date"]:
            try:
                date = parsedate_to_datetime(msg["Date"])
            except Exception:
                pass

        return EmailMetadata(
            subject=msg.get("Subject", ""),
            mail_from=msg.get("From", ""),
            mail_to=[addr for _, addr in getaddresses(msg.get_all("To", []))],
            mail_cc=[addr for _, addr in getaddresses(msg.get_all("Cc", []))],
            date=date,
            message_id=msg.get("Message-ID"),
        )

    def _extract_attachments(self, msg) -> list[Attachment]:
        attachments = []
        for part in msg.walk():
            if part.get_content_disposition() != "attachment":
                continue
            filename = self._decode_filename(part.get_filename())
            if not filename:
                continue
            content = part.get_payload(decode=True) or b""
            att = Attachment(
                filename=filename,
                content_type=part.get_content_type(),
                content=content,
            )
            attachments.append(att)
        return attachments

    def _decode_filename(self, raw: str | None) -> str | None:
        if not raw:
            return None
        parts = decode_header(raw)
        return "".join(
            chunk.decode(enc or "utf-8") if isinstance(chunk, bytes) else chunk
            for chunk, enc in parts
        )

    def enrich(self, att: Attachment) -> None:
        att.info = self._enrich(att)
        if att.is_excel() and att.info is not None:
            wb = openpyxl.load_workbook(io.BytesIO(att.content), data_only=True)
            att.info.parsed_data = get_processor(att).process(wb)
            wb.close()

    def _enrich(self, att: Attachment) -> ExcelInfo | PDFInfo | None:
        if att.is_excel():
            return self._enrich_excel(att.content)
        if att.is_pdf():
            return self._enrich_pdf(att.content)
        return None

    def _enrich_excel(self, content: bytes) -> ExcelInfo:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = wb.sheetnames
        row_count: dict[str, int] = {}
        col_count: dict[str, int] = {}
        headers: dict[str, list[str]] = {}

        for name in sheets:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            row_count[name] = len(rows)
            col_count[name] = max((len(r) for r in rows), default=0)
            first_row = rows[0] if rows else []
            headers[name] = [str(c) if c is not None else "" for c in first_row]

        first_sheet_text = self._sheet_to_text(wb[sheets[0]]) if sheets else None
        wb.close()
        info = ExcelInfo(
            sheets=sheets,
            row_count=row_count,
            col_count=col_count,
            headers=headers,
            sheet_text=first_sheet_text,
        )
        # info.llm_info = self._enrich_excel_with_llm(info)  # TODO: activar cuando el LLM esté configurado
        return info

    def _sheet_to_text(self, ws) -> str:
        lines = []
        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            lines.append("\t".join("" if cell is None else str(cell) for cell in row))
        return "\n".join(lines)

    def _enrich_excel_with_llm(self, info: ExcelInfo) -> ExcelLLMInfo:
        # TODO: construir prompt a partir de info.sheets, info.headers, etc.
        # TODO: llamar al LLM y parsear la respuesta
        raise NotImplementedError

    def _enrich_pdf(self, content: bytes) -> PDFInfo:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            page_count = len(pdf.pages)
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
                if len(text) >= 500:
                    break
        return PDFInfo(
            page_count=page_count,
            text_preview=text[:500] if text else None,
        )
